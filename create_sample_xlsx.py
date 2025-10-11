#!/usr/bin/env python3
"""
Script to create a sample XLSX file with the correct structure for database credentials.
This is a one-time helper script to generate the template.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Database Credentials"

# Add headers with styling
headers = ["database_name", "username", "password", "host"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Add sample data rows
sample_data = [
    ["myapp_db", "myapp_user", "secure_password_123", "localhost"],
    ["analytics_db", "analytics_user", "another_password_456", "localhost"],
    ["staging_db", "staging_user", "staging_pass_789", "localhost"],
]

for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Adjust column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15

# Save the file
wb.save("database_credentials_SAMPLE.xlsx")
print("✅ Sample XLSX file created: database_credentials_SAMPLE.xlsx")
print("\nStructure:")
print("  Column A: database_name (name of the database)")
print("  Column B: username (MySQL username for this database)")
print("  Column C: password (MySQL password for this user)")
print("  Column D: host (usually 'localhost' or IP address)")
print("\nFill this file with your actual database credentials and save as 'database_credentials.xlsx'")
